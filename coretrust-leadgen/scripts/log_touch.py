#!/usr/bin/env python3
"""
Log a touch against a lead in CoreTrust_Master_Members.xlsx from the command
line -- the same "read then update-or-add" pattern the Log Touch Copilot Studio
flow uses, for when you're working the file locally without Copilot Studio
open (offline, testing, or before the flows are built).

Usage:
    python3 log_touch.py --lead CT-00001 --channel Email
    python3 log_touch.py --lead CT-00042 --channel Call --reply --outcome replied
    python3 log_touch.py --lead CT-25442 --channel Email --meeting --outcome meeting

If the lead has no TOUCHPOINTS row yet, one is added (pulling Company_Name,
contact, tier, priority, supplier from MASTER MEMBERS) with Cadence_Step 1 and
the CADENCE-driven follow-up formulas already wired in. If it does, Total_Touches
and the channel-specific counter increment, Last_Touch_Date/Channel update, and
Cadence_Step advances by one (capped at 8, the length of CADENCE).
"""
import argparse
import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DEFAULT_FILE = Path(__file__).parent.parent / "data" / "CoreTrust_Master_Members.xlsx"
CHANNEL_COL = {"Email": "Emails_Sent", "Call": "Calls_Made", "LinkedIn": "LinkedIn_Touches"}


def sheet_index(ws):
    headers = [c.value for c in ws[1]]
    return {h: i + 1 for i, h in enumerate(headers)}  # 1-based column numbers


def find_row(ws, idx, key_col, key_val):
    col = idx[key_col]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=col).value == key_val:
            return r
    return None


def expand_table(ws, table_name, last_row):
    tbl = ws.tables.get(table_name)
    if tbl is None:
        return
    start, _, end_col_letter, _ = _parse_ref(tbl.ref)
    ncols = ws.max_column
    tbl.ref = f"{start}:{get_column_letter(ncols)}{last_row}"
    if tbl.autoFilter is not None:
        tbl.autoFilter.ref = tbl.ref


def _parse_ref(ref):
    start, end = ref.split(":")
    import re
    m = re.match(r"([A-Z]+)(\d+)", end)
    return start, None, m.group(1), int(m.group(2))


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--lead", required=True, help="Lead_ID, e.g. CT-00001")
    ap.add_argument("--channel", required=True, choices=["Email", "Call", "LinkedIn"])
    ap.add_argument("--reply", action="store_true", help="mark Reply_Received = Yes")
    ap.add_argument("--meeting", action="store_true", help="mark Meeting_Booked = Yes")
    ap.add_argument("--outcome", default=None, help="contacted / replied / meeting / disqualified")
    ap.add_argument("--rep", default="(rep)", help="Assigned_Rep, only used when creating a new row")
    ap.add_argument("--file", default=str(DEFAULT_FILE))
    args = ap.parse_args()

    path = Path(args.file)
    wb = load_workbook(path)
    master = wb["MASTER MEMBERS"]
    touch = wb["TOUCHPOINTS"]
    m_idx = sheet_index(master)
    t_idx = sheet_index(touch)
    today = datetime.date.today()

    m_row = find_row(master, m_idx, "Lead_ID", args.lead)
    if m_row is None:
        raise SystemExit(f"Lead_ID {args.lead} not found in MASTER MEMBERS")

    def mval(col):
        return master.cell(row=m_row, column=m_idx[col]).value

    t_row = find_row(touch, t_idx, "Lead_ID", args.lead)
    if t_row is None:
        t_row = touch.max_row + 1
        touch.cell(row=t_row, column=t_idx["Lead_ID"], value=args.lead)
        touch.cell(row=t_row, column=t_idx["Company_Name"], value=mval("Company_Name"))
        touch.cell(row=t_row, column=t_idx["Contact_First"], value=mval("Contact_First"))
        touch.cell(row=t_row, column=t_idx["Contact_Last"], value=mval("Contact_Last"))
        touch.cell(row=t_row, column=t_idx["Contact_Title"], value=mval("Contact_Title"))
        touch.cell(row=t_row, column=t_idx["Contact_Email"], value=mval("Contact_Email"))
        touch.cell(row=t_row, column=t_idx["Assigned_Rep"], value=args.rep)
        touch.cell(row=t_row, column=t_idx["Fit_Tier"], value=mval("Fit_Tier"))
        touch.cell(row=t_row, column=t_idx["Priority"], value=mval("Priority"))
        touch.cell(row=t_row, column=t_idx["Spend_Bucket"], value=mval("Spend_Bucket"))
        touch.cell(row=t_row, column=t_idx["Suggested_Supplier"], value=mval("Suggested_Supplier"))
        touch.cell(row=t_row, column=t_idx["Cadence_Step"], value=1)
        touch.cell(row=t_row, column=t_idx["Total_Touches"], value=0)
        touch.cell(row=t_row, column=t_idx["Emails_Sent"], value=0)
        touch.cell(row=t_row, column=t_idx["Calls_Made"], value=0)
        touch.cell(row=t_row, column=t_idx["LinkedIn_Touches"], value=0)
        touch.cell(row=t_row, column=t_idx["Reply_Received"], value="No")
        touch.cell(row=t_row, column=t_idx["Meeting_Booked"], value="No")
        touch.cell(row=t_row, column=t_idx["Outcome"], value="contacted")
        touch.cell(row=t_row, column=t_idx["Cadence_Status"], value="Active")
        Q = get_column_letter(t_idx["Last_Touch_Date"])
        L = get_column_letter(t_idx["Cadence_Step"])
        S = get_column_letter(t_idx["Days_Since_Last"])
        T = get_column_letter(t_idx["Next_Touch_Due"])
        touch.cell(row=t_row, column=t_idx["Days_Since_Last"],
                   value=f'=IF({Q}{t_row}="","",TODAY()-{Q}{t_row})')
        touch.cell(row=t_row, column=t_idx["Next_Touch_Due"],
                   value=f'=IF({Q}{t_row}="","",{Q}{t_row}+VLOOKUP({L}{t_row},CADENCE!$A:$B,2,FALSE))')
        touch.cell(row=t_row, column=t_idx["Follow_Up_Flag"],
                   value=f'=IF({T}{t_row}="","",IF(TODAY()>={T}{t_row},"FOLLOW UP","ok"))')
        touch.cell(row=t_row, column=t_idx["SLA_Status"],
                   value=f'=IF({S}{t_row}="","",IF({S}{t_row}>14,"OVERDUE",IF({S}{t_row}>4,"DUE SOON","ON TRACK")))')
        expand_table(touch, "TouchTable", t_row)
        print(f"Added new TOUCHPOINTS row for {args.lead} ({mval('Company_Name')})")
    else:
        step = touch.cell(row=t_row, column=t_idx["Cadence_Step"]).value or 1
        touch.cell(row=t_row, column=t_idx["Cadence_Step"], value=min(int(step) + 1, 8))

    total = touch.cell(row=t_row, column=t_idx["Total_Touches"]).value or 0
    touch.cell(row=t_row, column=t_idx["Total_Touches"], value=int(total) + 1)
    ch_col = CHANNEL_COL[args.channel]
    n = touch.cell(row=t_row, column=t_idx[ch_col]).value or 0
    touch.cell(row=t_row, column=t_idx[ch_col], value=int(n) + 1)
    touch.cell(row=t_row, column=t_idx["Last_Touch_Date"], value=today)
    touch.cell(row=t_row, column=t_idx["Last_Touch_Channel"], value=args.channel)
    if args.reply:
        touch.cell(row=t_row, column=t_idx["Reply_Received"], value="Yes")
    if args.meeting:
        touch.cell(row=t_row, column=t_idx["Meeting_Booked"], value="Yes")
    if args.outcome:
        touch.cell(row=t_row, column=t_idx["Outcome"], value=args.outcome)

    if str(master.cell(row=m_row, column=m_idx["Record_Status"]).value or "") in ("", "Ready to Call", "New", "Enriching"):
        master.cell(row=m_row, column=m_idx["Record_Status"], value="In Cadence")
        master.cell(row=m_row, column=m_idx["Updated_By"], value="Rep (log_touch.py)")
        master.cell(row=m_row, column=m_idx["Last_Updated"], value=today)

    wb.save(path)
    print(f"Logged a {args.channel} touch for {args.lead} ({mval('Company_Name')}). Saved {path}")


if __name__ == "__main__":
    main()
