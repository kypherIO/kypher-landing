#!/usr/bin/env python3
"""
Save BANTC qualification markers for one lead from the command line -- the
same job the Save Qualification Copilot Studio flow does, for working the
file locally. Writes the five markers (TMS_In_Use, Under_Contract,
Contract_Status, Capacity_Source, Private_Fleet), recomputes BANT_Budget /
BANT_Authority / BANT_Need / BANT_Timeline / BANTC_Status from whatever is
now on the row, and -- this is the "trigger a meeting with category SME"
step -- appends a row to SME HANDOFF the moment BANTC_Status flips to
"Qualified - Ready for SME", pulling the right SME from the SME ROUTING
table by the lead's Category.

Usage:
    python3 save_qualification.py --lead CT-00001 \
        --tms "Optimized TMS" --contract-status "Renewal <6mo" \
        --capacity-source "Broker/Spot" --private-fleet No --under-contract Yes \
        --confirmed-freight-spend 4200000

Also applies the SDR Playbook's "Retry, Nurture, and Recycle Rules" (a
disposition reason on TOUCHPOINTS, never a deleted record) via --disposition:

    python3 save_qualification.py --lead CT-00001 --disposition "Nonbuyer - suppress"
    python3 save_qualification.py --lead CT-00001 --disposition "New trigger - recycle"
    python3 save_qualification.py --lead CT-00001 --disposition "Timing not right - dated follow-up" \
        --followup-date 2026-11-01
"""
import argparse
import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DEFAULT_FILE = Path(__file__).parent.parent / "data" / "CoreTrust_Master_Members.xlsx"

MARKER_ARGS = {
    "tms": "TMS_In_Use",
    "under_contract": "Under_Contract",
    "contract_status": "Contract_Status",
    "capacity_source": "Capacity_Source",
    "private_fleet": "Private_Fleet",
    "confirmed_freight_spend": "Confirmed_Freight_Spend",
}

AUTHORITY_TITLES = ("vp", "vice president", "director", "chief", "svp", "evp", "president",
                    "head of", "cfo", "coo", "ceo", "cpo", "owner", "partner")

DISPOSITION_REASONS = [
    "Nonbuyer - suppress", "No response after 8 touches - nurture",
    "Timing not right - dated follow-up", "New trigger - recycle",
    "Disqualified - BANTC gate failed", "Other",
]


def norm_str(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def bant_budget_of(est, confirmed):
    if confirmed and float(confirmed) > 0:
        return "Yes"
    if est and float(est) >= 1_000_000:
        return "Yes"
    if est and float(est) > 0:
        return "Unconfirmed"
    return ""


def bant_authority_of(title, email):
    t = norm_str(title).lower()
    if not t:
        return ""
    return "Yes" if any(k in t for k in AUTHORITY_TITLES) and norm_str(email) else "Unconfirmed"


def bant_need_of(tms, private_fleet, capacity_source, under_contract):
    if not any(norm_str(v) for v in (tms, private_fleet, capacity_source, under_contract)):
        return ""
    pain = (norm_str(tms) == "Manual / No TMS" or norm_str(capacity_source) in ("Broker/Spot", "Mixed")
            or norm_str(under_contract).lower() == "no")
    return "Yes" if pain else "No"


def bant_timeline_of(contract_status):
    s = norm_str(contract_status)
    if not s:
        return ""
    return "Yes" if s in ("None", "Renewal <6mo") else "No"


def bantc_status_of(budget, authority, need, timeline):
    yes = sum(1 for v in (budget, authority, need, timeline) if v == "Yes")
    if budget == "Yes" and authority == "Yes" and yes >= 3:
        return "Qualified - Ready for SME"
    if any((budget, authority, need, timeline)):
        return "In Progress"
    return "Not Started"


def sheet_index(ws):
    return {c.value: i + 1 for i, c in enumerate(ws[1])}


def find_row(ws, idx, key_col, key_val):
    col = idx[key_col]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=col).value == key_val:
            return r
    return None


def expand_table(ws, table_name, last_row):
    tbl = ws.tables.get(table_name)
    if tbl is None:
        return
    start = tbl.ref.split(":")[0]
    tbl.ref = f"{start}:{get_column_letter(ws.max_column)}{last_row}"
    if tbl.autoFilter is not None:
        tbl.autoFilter.ref = tbl.ref


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--lead", required=True)
    ap.add_argument("--tms", choices=["Manual / No TMS", "Basic TMS", "Optimized TMS"])
    ap.add_argument("--under-contract", choices=["Yes", "No"])
    ap.add_argument("--contract-status", choices=["None", "Month-to-month", "Renewal <6mo", "Locked"])
    ap.add_argument("--capacity-source", choices=["Broker/Spot", "Mixed", "Direct/Contracted", "Private Fleet"])
    ap.add_argument("--private-fleet", choices=["Yes", "No"])
    ap.add_argument("--confirmed-freight-spend", type=float)
    ap.add_argument("--disposition", choices=DISPOSITION_REASONS,
                     help="Apply a Retry/Nurture/Recycle disposition. Never deletes the record.")
    ap.add_argument("--followup-date", help="YYYY-MM-DD, used with --disposition "
                     "'Timing not right - dated follow-up'.")
    ap.add_argument("--file", default=str(DEFAULT_FILE))
    args = ap.parse_args()

    path = Path(args.file)
    wb = load_workbook(path)
    master = wb["MASTER MEMBERS"]
    m_idx = sheet_index(master)
    today = datetime.date.today()

    m_row = find_row(master, m_idx, "Lead_ID", args.lead)
    if m_row is None:
        raise SystemExit(f"Lead_ID {args.lead} not found in MASTER MEMBERS")

    def mget(col):
        return master.cell(row=m_row, column=m_idx[col]).value

    def mset(col, val):
        if val is not None:
            master.cell(row=m_row, column=m_idx[col], value=val)

    given = {
        "TMS_In_Use": args.tms, "Under_Contract": args.under_contract,
        "Contract_Status": args.contract_status, "Capacity_Source": args.capacity_source,
        "Private_Fleet": args.private_fleet, "Confirmed_Freight_Spend": args.confirmed_freight_spend,
    }
    for col, val in given.items():
        mset(col, val)

    budget = bant_budget_of(mget("Est_Freight_Spend"), mget("Confirmed_Freight_Spend"))
    authority = bant_authority_of(mget("Contact_Title"), mget("Contact_Email"))
    need = bant_need_of(mget("TMS_In_Use"), mget("Private_Fleet"), mget("Capacity_Source"), mget("Under_Contract"))
    timeline = bant_timeline_of(mget("Contract_Status"))
    status = bantc_status_of(budget, authority, need, timeline)

    mset("BANT_Budget", budget)
    mset("BANT_Authority", authority)
    mset("BANT_Need", need)
    mset("BANT_Timeline", timeline)
    mset("BANTC_Status", status)
    mset("Updated_By", "Rep (save_qualification.py)")
    mset("Last_Updated", today)

    markers_set = sum(1 for c in ("TMS_In_Use", "Under_Contract", "Contract_Status", "Capacity_Source",
                                    "Private_Fleet", "Confirmed_Freight_Spend") if norm_str(mget(c)))
    if markers_set >= 3 and norm_str(mget("Private_Fleet")).lower() != "yes":
        mset("Qualified", "Yes")
        mset("Record_Status", "Qualified")

    print(f"{args.lead}: BANT Budget={budget} Authority={authority} Need={need} Timeline={timeline} "
          f"-> {status}")

    if status == "Qualified - Ready for SME":
        routing = wb["SME ROUTING"]
        r_idx = sheet_index(routing)
        category = mget("Category")
        sme_name = sme_email = ""
        for r in range(2, routing.max_row + 1):
            if routing.cell(row=r, column=r_idx["Category"]).value == category:
                sme_name = routing.cell(row=r, column=r_idx["SME_Name"]).value or ""
                sme_email = routing.cell(row=r, column=r_idx["SME_Email"]).value or ""
                break

        handoff = wb["SME HANDOFF"]
        h_idx = sheet_index(handoff)
        h_row = find_row(handoff, h_idx, "Lead_ID", args.lead)
        if h_row is None:
            h_row = handoff.max_row + 1
            handoff.cell(row=h_row, column=h_idx["Lead_ID"], value=args.lead)
            handoff.cell(row=h_row, column=h_idx["Company_Name"], value=mget("Company_Name"))
            handoff.cell(row=h_row, column=h_idx["Category"], value=category)
            handoff.cell(row=h_row, column=h_idx["BANTC_Status"], value=status)
            handoff.cell(row=h_row, column=h_idx["Fit_Tier"], value=mget("Fit_Tier"))
            handoff.cell(row=h_row, column=h_idx["Est_Freight_Spend"], value=mget("Est_Freight_Spend"))
            handoff.cell(row=h_row, column=h_idx["Contact_First"], value=mget("Contact_First"))
            handoff.cell(row=h_row, column=h_idx["Contact_Last"], value=mget("Contact_Last"))
            handoff.cell(row=h_row, column=h_idx["Contact_Title"], value=mget("Contact_Title"))
            handoff.cell(row=h_row, column=h_idx["Contact_Email"], value=mget("Contact_Email"))
            handoff.cell(row=h_row, column=h_idx["CT_Account_Manager"], value=mget("CT_Account_Manager"))
            handoff.cell(row=h_row, column=h_idx["SME_Name"], value=sme_name)
            handoff.cell(row=h_row, column=h_idx["SME_Email"], value=sme_email)
            handoff.cell(row=h_row, column=h_idx["Meeting_Requested_Date"], value=today)
            handoff.cell(row=h_row, column=h_idx["Meeting_Status"], value="Requested")
            expand_table(handoff, "SMEHandoffTable", h_row)
            print(f"  -> BANTC gate cleared. Added to SME HANDOFF, routed to {sme_name or '(no SME set for '+str(category)+' in SME ROUTING)'}.")
        else:
            handoff.cell(row=h_row, column=h_idx["BANTC_Status"], value=status)
            print(f"  -> Already on SME HANDOFF (row {h_row}); status refreshed.")

    if args.disposition:
        touch = wb["TOUCHPOINTS"]
        t_idx = sheet_index(touch)
        t_row = find_row(touch, t_idx, "Lead_ID", args.lead)
        if t_row is None:
            # Applying a disposition to a lead with no TOUCHPOINTS row yet (e.g.
            # disqualifying before a single touch went out) -- add a minimal row
            # rather than silently doing nothing.
            t_row = touch.max_row + 1
            touch.cell(row=t_row, column=t_idx["Lead_ID"], value=args.lead)
            touch.cell(row=t_row, column=t_idx["Company_Name"], value=mget("Company_Name"))
            touch.cell(row=t_row, column=t_idx["Fit_Tier"], value=mget("Fit_Tier"))
            touch.cell(row=t_row, column=t_idx["Cadence_Step"], value=1)
            touch.cell(row=t_row, column=t_idx["Total_Touches"], value=0)
            expand_table(touch, "TouchTable", t_row)

        touch.cell(row=t_row, column=t_idx["Disposition_Reason"], value=args.disposition)

        if args.disposition == "Nonbuyer - suppress":
            touch.cell(row=t_row, column=t_idx["Cadence_Status"], value="Suppressed")
            mset("Record_Status", "Disqualified")
            mset("Qualified", "No")
        elif args.disposition == "Disqualified - BANTC gate failed":
            touch.cell(row=t_row, column=t_idx["Cadence_Status"], value="Suppressed")
            mset("Record_Status", "Disqualified")
            mset("Qualified", "No")
        elif args.disposition == "No response after 8 touches - nurture":
            touch.cell(row=t_row, column=t_idx["Cadence_Status"], value="Nurture")
        elif args.disposition == "New trigger - recycle":
            # Re-enter the cadence from the top, per the playbook: reset the step,
            # go back to Active. The record was never deleted, so nothing else to undo.
            touch.cell(row=t_row, column=t_idx["Cadence_Status"], value="Active")
            touch.cell(row=t_row, column=t_idx["Cadence_Step"], value=1)
        elif args.disposition == "Timing not right - dated follow-up":
            touch.cell(row=t_row, column=t_idx["Cadence_Status"], value="Nurture")
            if args.followup_date:
                touch.cell(row=t_row, column=t_idx["Scheduled_Followup_Date"],
                           value=datetime.date.fromisoformat(args.followup_date))

        print(f"  -> Disposition set: {args.disposition}"
              + (f", follow-up scheduled {args.followup_date}" if args.followup_date else ""))

    wb.save(path)
    print(f"Saved {path}")


if __name__ == "__main__":
    main()
