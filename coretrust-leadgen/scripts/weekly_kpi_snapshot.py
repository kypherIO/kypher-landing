#!/usr/bin/env python3
"""
Append this week's demand-gen numbers to the "Weekly Demand-Gen History" table
on the KPI DASHBOARD tab of CoreTrust_Master_Members.xlsx.

The live tiles at the top of KPI DASHBOARD are Excel formulas (COUNTIF/SUMIF
over MasterTable and TouchTable) -- they always show today's numbers when the
workbook is open, but Excel formulas don't keep history. This script computes
the same numbers directly from the sheet data (no dependency on Excel having
recalculated) and writes one new row of plain values, so the history survives
even as the live totals keep moving. Run it weekly -- by hand, from Task
Scheduler / a OneDrive-synced cron, or as the last step of the Nightly
Enrichment Copilot Studio flow.

Usage:
    python3 weekly_kpi_snapshot.py
    python3 weekly_kpi_snapshot.py --week-of 2026-09-07 --file "path/to/CoreTrust_Master_Members.xlsx"
"""
import argparse
import datetime
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_FILE = Path(__file__).parent.parent / "data" / "CoreTrust_Master_Members.xlsx"
HEADERS = ["Week_Of", "Total_Members", "Freight_Relevant", "Tier_A", "Tier_B", "Verified_Contacts",
           "New_Or_Enriching", "Touches_Sent", "Meetings_Booked", "Qualified_Opps", "Addressable_Freight"]


def sheet_rows(ws, headers_row=1):
    headers = [c.value for c in ws[headers_row]]
    idx = {h: i for i, h in enumerate(headers)}
    for row in ws.iter_rows(min_row=headers_row + 1, values_only=True):
        if row[idx.get("Lead_ID", 0)] is None:
            continue
        yield idx, row


def compute_stats(wb):
    master = wb["MASTER MEMBERS"]
    total = freight = tier_a = tier_b = verified = new_or_enriching = qualified = 0
    addressable = 0.0
    for idx, row in sheet_rows(master):
        total += 1
        fr = str(row[idx["Freight_Relevant"]] or "").lower()
        if fr == "yes":
            freight += 1
            est = row[idx["Est_Freight_Spend"]]
            if est:
                addressable += float(est)
        tier = row[idx["Fit_Tier"]]
        if tier == "A":
            tier_a += 1
        elif tier == "B":
            tier_b += 1
        email = row[idx["Contact_Email"]]
        if email and "@" in str(email):
            verified += 1
        status = row[idx["Record_Status"]]
        if status in ("New", "Enriching"):
            new_or_enriching += 1
        q = str(row[idx["Qualified"]] or "").lower()
        if q == "yes":
            qualified += 1

    touch = wb["TOUCHPOINTS"]
    touches = meetings = 0
    for idx, row in sheet_rows(touch):
        tt = row[idx["Total_Touches"]]
        if tt:
            touches += int(tt)
        mb = str(row[idx["Meeting_Booked"]] or "").lower()
        if mb == "yes":
            meetings += 1

    return {
        "Total_Members": total, "Freight_Relevant": freight, "Tier_A": tier_a, "Tier_B": tier_b,
        "Verified_Contacts": verified, "New_Or_Enriching": new_or_enriching, "Touches_Sent": touches,
        "Meetings_Booked": meetings, "Qualified_Opps": qualified, "Addressable_Freight": round(addressable, 2),
    }


def find_history_block(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3):
        for cell in row:
            if cell.value == "Week_Of":
                header_row = cell.row
                header_col = cell.column
                last_row = header_row
                r = header_row + 1
                while ws.cell(row=r, column=header_col).value is not None:
                    last_row = r
                    r += 1
                return header_row, header_col, last_row
    raise RuntimeError('Could not find the "Week_Of" header on KPI DASHBOARD')


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--file", default=str(DEFAULT_FILE))
    ap.add_argument("--week-of", default=None, help="YYYY-MM-DD, defaults to today")
    args = ap.parse_args()

    path = Path(args.file)
    wb = load_workbook(path)
    stats = compute_stats(wb)

    ws = wb["KPI DASHBOARD"]
    header_row, header_col, last_row = find_history_block(ws)
    new_row = last_row + 1

    week_of = datetime.date.fromisoformat(args.week_of) if args.week_of else datetime.date.today()
    values = [week_of] + [stats[h] for h in HEADERS[1:]]
    for i, v in enumerate(values):
        ws.cell(row=new_row, column=header_col + i, value=v)

    wb.save(path)
    print(f"Appended a KPI snapshot for {week_of.isoformat()} to KPI DASHBOARD row {new_row}:")
    for h, v in zip(HEADERS, values):
        print(f"  {h}: {v}")
    print(f"Saved {path}")


if __name__ == "__main__":
    main()
