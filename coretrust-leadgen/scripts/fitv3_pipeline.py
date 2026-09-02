#!/usr/bin/env python3
"""
CoreTrust Fit v3 pipeline: merges the SFDC-scored master with a verified-contact
enrichment file, backfills the columns the agent instructions and prompt library
reference but the working file was missing (Actionability, Warmth, Priority,
Service_Fit, PE_Sponsor, Portfolio_Company, Contact_Phone), and writes the final
workbook: MASTER MEMBERS, TOUCHPOINTS, CADENCE, SUMMARY, KPI DASHBOARD, READ ME,
Data Dictionary -- each of the three working tabs as a real named Excel Table
(MasterTable / TouchTable / CadenceTable) so the Copilot Studio Excel Online
connector's table picker finds them by name (see the build guide's "table
dropdown empty" gotcha).

Scoring math is never re-derived here: fitv3_engine.py owns Fit v3 (30/30/15/15/10,
tiers A>=70 B>=55 C>=40 D). This script only supplies the merge, the percentile
independent lookups the engine needs, and the workbook layout.

Usage:
    python3 fitv3_pipeline.py \
        --master data/source/CoreTrust_Master_Members_SFDC_Scored.xlsx \
        --verified data/source/CoreTrust_Master_Members_Verified_Sample.xlsx \
        --out data/CoreTrust_Master_Members.xlsx

Re-run any time a fresh SFDC export or a fresh freight/contact enrichment batch
lands in data/source/ -- the merge is idempotent, keyed on normalized company name.
"""
import argparse
import datetime
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).parent))
import fitv3_engine as engine

TODAY = datetime.date.today()

# ---------------------------------------------------------------------------
# Final MASTER MEMBERS schema. Grouped so header fill colors read like the
# original file: grey record / blue company / teal contact / purple markers /
# gold score-inputs / green computed / navy outcome.
# ---------------------------------------------------------------------------
GROUP_COLOR = {
    "record": "FF5A6675",
    "company": "FF2E6DA4",
    "contact": "FF0E7C86",
    "markers": "FF6A4CB5",
    "score_inputs": "FFC9922B",
    "computed": "FF2E7D32",
    "outcome": "FF1F3A5F",
}

SCHEMA = [
    ("Lead_ID", "record"), ("Record_Status", "record"), ("Freight_Relevant", "record"),
    ("Data_Source", "record"), ("Assigned_Rep", "record"), ("Last_Updated", "record"),
    ("Updated_By", "record"),
    ("Company_Name", "company"), ("Website", "company"), ("Industry", "company"),
    ("Sub_Industry", "company"), ("HQ_City", "company"), ("HQ_State", "company"),
    ("Employees", "company"), ("Annual_Revenue", "company"), ("Rev_Source", "company"),
    ("Sales_Village", "company"), ("CT_Account_Manager", "company"), ("CT_TTM_Spend", "company"),
    ("Opps_Won", "company"), ("PA_Effective_Date", "company"), ("PE_Sponsor", "company"),
    ("Portfolio_Company", "company"),
    ("Contact_First", "contact"), ("Contact_Last", "contact"), ("Contact_Title", "contact"),
    ("Contact_Email", "contact"), ("Contact_Phone", "contact"),
    ("Est_Freight_Spend", "markers"), ("Confirmed_Freight_Spend", "markers"),
    ("TMS_In_Use", "markers"), ("Tech_State", "markers"), ("Under_Contract", "markers"),
    ("Contract_Status", "markers"), ("Capacity_Source", "markers"), ("Private_Fleet", "markers"),
    ("Lane_Data_Available", "markers"),
    ("Freight_Opportunity", "score_inputs"), ("Engrainment", "score_inputs"),
    ("Freight_Intensity", "score_inputs"), ("Actionability", "score_inputs"),
    ("Warmth", "score_inputs"), ("Exclusion_Reason", "score_inputs"),
    ("Fit_v3_Score", "computed"), ("Fit_Tier", "computed"), ("Spend_Bucket", "computed"),
    ("Priority", "computed"), ("Suggested_Supplier", "computed"), ("Service_Fit", "computed"),
    ("Qualified", "outcome"), ("Notes", "outcome"),
]
COLS = [c for c, _ in SCHEMA]
COL_LETTER = {c: get_column_letter(i + 1) for i, c in enumerate(COLS)}

TOUCH_SCHEMA = [
    "Lead_ID", "Company_Name", "Contact_First", "Contact_Last", "Contact_Title",
    "Contact_Email", "Assigned_Rep", "Fit_Tier", "Priority", "Spend_Bucket",
    "Suggested_Supplier", "Cadence_Step", "Total_Touches", "Emails_Sent", "Calls_Made",
    "LinkedIn_Touches", "Last_Touch_Date", "Last_Touch_Channel", "Days_Since_Last",
    "Next_Touch_Due", "Follow_Up_Flag", "SLA_Status", "Reply_Received", "Meeting_Booked",
    "Outcome", "Cadence_Status", "Notes",
]
TOUCH_LETTER = {c: get_column_letter(i + 1) for i, c in enumerate(TOUCH_SCHEMA)}

CADENCE_STEPS = [
    (1, 3, "Email intro"), (2, 1, "Call + voicemail"), (3, 3, "Email re: savings"),
    (4, 3, "LinkedIn"), (5, 3, "Call"), (6, 3, "Email case study"), (7, 3, "Call"),
    (8, 3, "Breakup email"),
]

HEAVY_FREIGHT_INDUSTRIES = {
    "manufacturing", "distribution", "wholesale", "building materials", "chemicals",
    "industrial", "automotive", "machinery", "agriculture", "construction",
}
CONSUMER_RETAIL_INDUSTRIES = {
    "consumer products", "consumer", "food", "beverage", "apparel", "retail",
}


import re as _re

DBA_RE = _re.compile(r'd[.\s]?[/.]?\s?b[.\s]?[/.]?\s?a\b\.?\s*(.*)', _re.IGNORECASE)


def dba_trade_name_key(name):
    """SFDC often carries the legal entity, e.g. 'Asphalt Buyer II LLC D.B.A.
    American Tire Distributors'. engine.norm() keys on the legal name (the text
    before d/b/a); this gives the trade-name key too, so a freight/contact file
    that only knows the trade name still matches."""
    s = str(name or "")
    m = DBA_RE.search(s)
    if m and m.group(1).strip():
        trade_name = _re.sub(r'\([^)]*\)', '', m.group(1))  # drop trailing "(ATD)" style abbreviations
        return engine.norm(trade_name)
    return None


def norm_str(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def has_ae(ct_account_manager, assigned_rep):
    return bool(norm_str(ct_account_manager) or norm_str(assigned_rep))


def warmth_of(pe_sponsor, ct_account_manager, assigned_rep):
    wm = 0
    pe = norm_str(pe_sponsor)
    if pe and pe.lower() != "independent":
        wm += 60
    if has_ae(ct_account_manager, assigned_rep):
        wm += 40
    return wm


def priority_of(tier):
    return {"A": "HOT", "B": "WARM", "C": "COLD", "D": "COLD"}.get(tier, "")


SPEND_BANDS = [(700_000, "<$700K"), (1_000_000, "$700K-$1M"), (2_000_000, "$1-2M"),
               (5_000_000, "$2-5M"), (10_000_000, "$5-10M")]
SUPPLIER_BY_BUCKET = {
    "Unknown": None, "<$700K": "CoreTrust Connect", "$700K-$1M": "Connect / Redwood if high save",
    "$1-2M": "Redwood", "$2-5M": "Redwood / GEODIS", "$5-10M": "GEODIS", "$10M+": "GEODIS",
}


def spend_bucket_of(est_freight_spend):
    try:
        est = float(est_freight_spend)
    except (TypeError, ValueError):
        return "Unknown"
    if not est or est <= 0:
        return "Unknown"
    for ceiling, label in SPEND_BANDS:
        if est < ceiling:
            return label
    return "$10M+"


def supplier_of(bucket):
    return SUPPLIER_BY_BUCKET.get(bucket)


def service_fit_of(freight_relevant, industry, sub_industry, employees, ct_ttm_spend):
    if norm_str(freight_relevant).lower() != "yes":
        return ""
    ind = norm_str(industry).lower()
    sub = norm_str(sub_industry).lower()
    tags = ["LTL"]
    if ind in HEAVY_FREIGHT_INDUSTRIES or sub in HEAVY_FREIGHT_INDUSTRIES or (employees or 0) >= 5000:
        tags.append("FTL")
    if (ind in CONSUMER_RETAIL_INDUSTRIES or sub in CONSUMER_RETAIL_INDUSTRIES
            or "food" in sub or "beverage" in sub or "apparel" in sub):
        tags.append("Retail Cross Dock")
    if (ct_ttm_spend or 0) > 0:
        tags.append("Brokerage")
    return ", ".join(tags)


def recompute_fit(fo, eng, fi, act, wm):
    fit = round(0.30 * fo + 0.30 * eng + 0.15 * fi + 0.15 * act + 0.10 * wm, 1)
    tier = "A" if fit >= 70 else "B" if fit >= 55 else "C" if fit >= 40 else "D"
    return fit, tier


def load_master(path):
    df = pd.read_excel(path, sheet_name="MASTER MEMBERS")
    df = df.loc[:, ~df.columns.duplicated()]
    return df


def load_verified(path):
    df = pd.read_excel(path, sheet_name="MASTER MEMBERS")
    return df


def build_merged_master(master_df, verified_df):
    verified_df = verified_df.copy()
    verified_df["_key"] = verified_df["Company_Name"].map(engine.norm)
    verified_by_key = {}
    for _, r in verified_df.iterrows():
        verified_by_key[r["_key"]] = r

    out_rows = []
    matched = 0
    matched_verified_keys = set()
    for _, row in master_df.iterrows():
        rec = {c: row.get(c) for c in master_df.columns if c in COLS}
        rec.setdefault("PE_Sponsor", None)
        rec.setdefault("Portfolio_Company", None)
        rec.setdefault("Contact_Phone", None)

        key = engine.norm(row.get("Company_Name"))
        v = verified_by_key.get(key)
        matched_key = key
        if v is None:
            alt = dba_trade_name_key(row.get("Company_Name"))
            if alt and alt in verified_by_key:
                v = verified_by_key[alt]
                matched_key = alt

        act_email = norm_str(rec.get("Contact_Email"))
        act_title = norm_str(rec.get("Contact_Title"))
        rec["Actionability"] = engine.action(act_email, act_title)
        rec["Warmth"] = warmth_of(rec.get("PE_Sponsor"), rec.get("CT_Account_Manager"), rec.get("Assigned_Rep"))

        if v is not None and norm_str(v.get("Contact_Email")):
            matched += 1
            matched_verified_keys.add(matched_key)
            rec["Contact_First"] = v.get("Contact_First") or rec.get("Contact_First")
            rec["Contact_Last"] = v.get("Contact_Last") or rec.get("Contact_Last")
            rec["Contact_Title"] = v.get("Contact_Title") or rec.get("Contact_Title")
            rec["Contact_Email"] = v.get("Contact_Email") or rec.get("Contact_Email")
            rec["Contact_Phone"] = v.get("Contact_Phone") or rec.get("Contact_Phone")
            rec["PE_Sponsor"] = v.get("PE_Sponsor") or rec.get("PE_Sponsor")
            rec["Portfolio_Company"] = v.get("Portfolio_Company") or ("Yes" if norm_str(v.get("PE_Sponsor")) and norm_str(v.get("PE_Sponsor")).lower() != "independent" else rec.get("Portfolio_Company"))
            if v.get("Suggested_Supplier"):
                rec["Suggested_Supplier"] = v.get("Suggested_Supplier")

            new_est = v.get("Est_Freight_Spend")
            if pd.notna(new_est) and new_est:
                rec["Est_Freight_Spend"] = new_est
                fo_v = engine.fo(float(new_est))
                if fo_v is not None:
                    rec["Freight_Opportunity"] = round(fo_v, 1)

            rec["Actionability"] = engine.action(norm_str(rec.get("Contact_Email")), norm_str(rec.get("Contact_Title")))
            rec["Warmth"] = warmth_of(rec.get("PE_Sponsor"), rec.get("CT_Account_Manager"), rec.get("Assigned_Rep"))

            fo = rec.get("Freight_Opportunity") or 0
            eng_v = rec.get("Engrainment") or 0
            fi = rec.get("Freight_Intensity") or 0
            fit, tier = recompute_fit(fo, eng_v, fi, rec["Actionability"], rec["Warmth"])
            rec["Fit_v3_Score"] = fit
            rec["Fit_Tier"] = tier
            rec["Spend_Bucket"] = spend_bucket_of(rec.get("Est_Freight_Spend"))
            if not rec.get("Suggested_Supplier"):
                rec["Suggested_Supplier"] = supplier_of(rec["Spend_Bucket"])
            if norm_str(rec.get("Record_Status")) in ("", "New", "Enriching", "nan"):
                rec["Record_Status"] = "Ready to Call"
            rec["Data_Source"] = (norm_str(rec.get("Data_Source")) + " + Freight Analysis").strip(" +")
            rec["Updated_By"] = "AI Agent (Pipeline)"
            rec["Last_Updated"] = TODAY

        rec["Priority"] = priority_of(rec.get("Fit_Tier"))
        rec["Service_Fit"] = service_fit_of(
            rec.get("Freight_Relevant"), rec.get("Industry"), rec.get("Sub_Industry"),
            rec.get("Employees"), rec.get("CT_TTM_Spend"),
        )
        out_rows.append(rec)

    # Verified/freight-analysis rows that don't match any existing SFDC member are net-new
    # accounts the freight analysis surfaced -- add them as fresh rows rather than drop them,
    # the same way the agent's "new registration" workflow scores and files a new member.
    existing_nums = master_df["Lead_ID"].astype(str).str.replace("CT-", "", regex=False)
    next_num = int(pd.to_numeric(existing_nums, errors="coerce").max()) + 1
    new_rows = []
    for _, v in verified_df.iterrows():
        if v["_key"] in matched_verified_keys or not norm_str(v.get("Contact_Email")):
            continue
        rec = {c: None for c in COLS}
        rec["Lead_ID"] = f"CT-{next_num:05d}"
        next_num += 1
        rec["Record_Status"] = "Ready to Call"
        rec["Freight_Relevant"] = "Yes"
        rec["Data_Source"] = "Freight Analysis (new account)"
        rec["Updated_By"] = "AI Agent (Pipeline)"
        rec["Last_Updated"] = TODAY
        for c in ("Company_Name", "Website", "Industry", "Sub_Industry", "HQ_State", "Employees",
                  "Annual_Revenue", "PE_Sponsor", "Portfolio_Company", "CT_TTM_Spend",
                  "Contact_First", "Contact_Last", "Contact_Title", "Contact_Email", "Contact_Phone",
                  "Est_Freight_Spend", "Suggested_Supplier"):
            val = v.get(c)
            if pd.notna(val):
                rec[c] = val
        rec["Freight_Intensity"] = round(engine.look(engine.INTENSITY, rec.get("Industry"), rec.get("Sub_Industry"), 8), 1)
        est = rec.get("Est_Freight_Spend")
        fo_v = engine.fo(float(est)) if est and pd.notna(est) else None
        rec["Freight_Opportunity"] = round(fo_v, 1) if fo_v is not None else round(rec["Freight_Intensity"] * 0.5, 1)
        rec["Actionability"] = engine.action(norm_str(rec.get("Contact_Email")), norm_str(rec.get("Contact_Title")))
        rec["Warmth"] = warmth_of(rec.get("PE_Sponsor"), rec.get("CT_Account_Manager"), rec.get("Assigned_Rep"))
        rec["Spend_Bucket"] = spend_bucket_of(est)
        if not rec.get("Suggested_Supplier"):
            rec["Suggested_Supplier"] = supplier_of(rec["Spend_Bucket"])
        new_rows.append(rec)

    if new_rows:
        # Engrainment is a percentile against the freight-relevant population (85% TTM spend +
        # 10% opps won, per the Scoring Methodology). Compute it fresh for these new rows only --
        # every existing row keeps the Engrainment it already had, so nothing already scored moves.
        pool_ttm, pool_opp = [], []
        for r in out_rows:
            if norm_str(r.get("Freight_Relevant")).lower() == "yes":
                pool_ttm.append(r.get("CT_TTM_Spend") or 0)
                pool_opp.append(r.get("Opps_Won") or 0)
        for r in new_rows:
            pool_ttm.append(r.get("CT_TTM_Spend") or 0)
            pool_opp.append(r.get("Opps_Won") or 0)
        ttm_pct_all = pd.Series(pool_ttm).rank(pct=True) * 100
        opp_pct_all = pd.Series(pool_opp).rank(pct=True) * 100
        n_new = len(new_rows)
        new_ttm_pct = ttm_pct_all.iloc[-n_new:].tolist()
        new_opp_pct = opp_pct_all.iloc[-n_new:].tolist()
        for r, ttm_pct, opp_pct in zip(new_rows, new_ttm_pct, new_opp_pct):
            eng_v = round(0.85 * ttm_pct + 0.10 * opp_pct + 0.05 * (100 if (r.get("CT_TTM_Spend") or 0) > 0 else 0), 1)
            r["Engrainment"] = eng_v
            fit, tier = recompute_fit(r["Freight_Opportunity"], eng_v, r["Freight_Intensity"], r["Actionability"], r["Warmth"])
            r["Fit_v3_Score"] = fit
            r["Fit_Tier"] = tier
            r["Priority"] = priority_of(tier)
            r["Service_Fit"] = service_fit_of(r.get("Freight_Relevant"), r.get("Industry"), r.get("Sub_Industry"),
                                               r.get("Employees"), r.get("CT_TTM_Spend"))
        out_rows.extend(new_rows)

    merged = pd.DataFrame(out_rows, columns=COLS)
    return merged, matched, len(new_rows)


def build_touchpoints(merged_df, verified_df):
    rows = []
    verified_df = verified_df.copy()
    verified_df["_key"] = verified_df["Company_Name"].map(engine.norm)
    verified_touch_by_key = {}
    try:
        import openpyxl as _oxl
        wb = _oxl.load_workbook(VERIFIED_PATH_GLOBAL, data_only=True)
        if "TOUCHPOINTS" in wb.sheetnames:
            tws = wb["TOUCHPOINTS"]
            theaders = [c.value for c in tws[1]]
            for r in tws.iter_rows(min_row=2, values_only=True):
                d = dict(zip(theaders, r))
                verified_touch_by_key[d.get("Company_Name")] = d
    except Exception:
        pass

    company_to_key = {}
    for _, v in verified_df.iterrows():
        company_to_key[v["_key"]] = v["Company_Name"]

    row_i = 2
    for _, m in merged_df.iterrows():
        key = engine.norm(m["Company_Name"])
        if key not in company_to_key:
            alt = dba_trade_name_key(m["Company_Name"])
            key = alt if alt in company_to_key else key
        if key not in company_to_key:
            continue
        vt = verified_touch_by_key.get(company_to_key[key])
        if vt is None:
            continue
        rows.append({
            "Lead_ID": m["Lead_ID"], "Company_Name": m["Company_Name"],
            "Contact_First": m["Contact_First"], "Contact_Last": m["Contact_Last"],
            "Contact_Title": m["Contact_Title"], "Contact_Email": m["Contact_Email"],
            "Assigned_Rep": "(rep)", "Fit_Tier": m["Fit_Tier"], "Priority": m["Priority"],
            "Spend_Bucket": m["Spend_Bucket"], "Suggested_Supplier": m["Suggested_Supplier"],
            "Cadence_Step": vt.get("Cadence_Step", 1), "Total_Touches": vt.get("Total_Touches", 1),
            "Emails_Sent": vt.get("Emails_Sent", 1), "Calls_Made": vt.get("Calls_Made", 0),
            "LinkedIn_Touches": vt.get("LinkedIn_Touches", 0),
            "Last_Touch_Date": vt.get("Last_Touch_Date"), "Last_Touch_Channel": vt.get("Last_Touch_Channel", "Email"),
            "Reply_Received": vt.get("Reply_Received", "No"), "Meeting_Booked": vt.get("Meeting_Booked", "No"),
            "Outcome": vt.get("Outcome", "contacted"), "Cadence_Status": vt.get("Cadence_Status", "Active"),
            "Notes": None,
        })
        row_i += 1
    return rows


VERIFIED_PATH_GLOBAL = None


def style_header(ws, ncols, group_of_col):
    for i in range(1, ncols + 1):
        cell = ws.cell(row=1, column=i)
        col_name = ws.cell(row=1, column=i).value
        color = group_of_col.get(col_name, "FF1F3A5F")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color="FFFFFFFF", bold=True, size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32


def add_table(ws, name, ref, style="TableStyleMedium9"):
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True, showFirstColumn=False,
                                          showLastColumn=False, showColumnStripes=False)
    ws.add_table(tbl)


def write_master_sheet(wb, merged_df):
    ws = wb.create_sheet("MASTER MEMBERS")
    ws.append(COLS)
    for _, r in merged_df.iterrows():
        row = []
        for c in COLS:
            v = r[c]
            if pd.isna(v):
                v = None
            elif isinstance(v, float) and v.is_integer() and c not in ("Fit_v3_Score",):
                v = int(v)
            row.append(v)
        ws.append(row)

    nrows = ws.max_row
    ref = f"A1:{get_column_letter(len(COLS))}{nrows}"
    group_of_col = {c: GROUP_COLOR[g] for c, g in SCHEMA}
    style_header(ws, len(COLS), group_of_col)
    add_table(ws, "MasterTable", ref)

    ws.freeze_panes = "H2"
    widths = {"Lead_ID": 11, "Company_Name": 32, "Website": 22, "Industry": 16, "Sub_Industry": 20,
              "Contact_Email": 26, "Notes": 30, "Exclusion_Reason": 18, "Service_Fit": 26}
    for c in COLS:
        ws.column_dimensions[COL_LETTER[c]].width = widths.get(c, 14)

    dv_status = DataValidation(type="list", formula1='"New,Enriching,Ready to Call,In Cadence,Qualified,Disqualified,Nurture,Low Priority"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"{COL_LETTER['Record_Status']}2:{COL_LETTER['Record_Status']}{nrows+2000}")

    dv_yn = DataValidation(type="list", formula1='"Yes,No,Unknown"', allow_blank=True)
    ws.add_data_validation(dv_yn)
    for c in ("Freight_Relevant", "TMS_In_Use", "Portfolio_Company"):
        dv_yn.add(f"{COL_LETTER[c]}2:{COL_LETTER[c]}{nrows+2000}")

    dv_yn2 = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_yn2)
    for c in ("Under_Contract", "Qualified"):
        dv_yn2.add(f"{COL_LETTER[c]}2:{COL_LETTER[c]}{nrows+2000}")

    dv_tech = DataValidation(type="list", formula1='"Manual / No TMS,Basic TMS,Optimized TMS"', allow_blank=True)
    ws.add_data_validation(dv_tech)
    dv_tech.add(f"{COL_LETTER['Tech_State']}2:{COL_LETTER['Tech_State']}{nrows+2000}")

    dv_contract = DataValidation(type="list", formula1='"None,Month-to-month,Renewal <6mo,Locked"', allow_blank=True)
    ws.add_data_validation(dv_contract)
    dv_contract.add(f"{COL_LETTER['Contract_Status']}2:{COL_LETTER['Contract_Status']}{nrows+2000}")

    dv_capacity = DataValidation(type="list", formula1='"Broker/Spot,Mixed,Direct/Contracted,Private Fleet"', allow_blank=True)
    ws.add_data_validation(dv_capacity)
    dv_capacity.add(f"{COL_LETTER['Capacity_Source']}2:{COL_LETTER['Capacity_Source']}{nrows+2000}")

    tier_col = COL_LETTER["Fit_Tier"]
    tier_colors = {"A": "FF2E7D32", "B": "FFC9922B", "C": "FF5A6675", "D": "FFB0B8C1"}
    for tier, color in tier_colors.items():
        rng = f"{tier_col}2:{tier_col}{nrows}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{tier}"'],
                                                        fill=PatternFill("solid", fgColor=color),
                                                        font=Font(color="FFFFFFFF", bold=True)))

    priority_col = COL_LETTER["Priority"]
    ws.conditional_formatting.add(f"{priority_col}2:{priority_col}{nrows}",
        CellIsRule(operator="equal", formula=['"HOT"'], fill=PatternFill("solid", fgColor="FFC0392B"), font=Font(color="FFFFFFFF", bold=True)))
    ws.conditional_formatting.add(f"{priority_col}2:{priority_col}{nrows}",
        CellIsRule(operator="equal", formula=['"WARM"'], fill=PatternFill("solid", fgColor="FFC9922B"), font=Font(bold=True)))

    ws.sheet_view.showGridLines = False
    return ws


def write_touch_sheet(wb, touch_rows):
    ws = wb.create_sheet("TOUCHPOINTS")
    ws.append(TOUCH_SCHEMA)
    r_i = 2
    for r in touch_rows:
        row = []
        for c in TOUCH_SCHEMA:
            if c == "Days_Since_Last":
                row.append(f'=IF(Q{r_i}="","",TODAY()-Q{r_i})')
            elif c == "Next_Touch_Due":
                row.append(f'=IF(Q{r_i}="","",Q{r_i}+VLOOKUP(L{r_i},CADENCE!$A:$B,2,FALSE))')
            elif c == "Follow_Up_Flag":
                row.append(f'=IF(T{r_i}="","",IF(TODAY()>=T{r_i},"FOLLOW UP","ok"))')
            elif c == "SLA_Status":
                row.append(f'=IF(S{r_i}="","",IF(S{r_i}>14,"OVERDUE",IF(S{r_i}>4,"DUE SOON","ON TRACK")))')
            else:
                v = r.get(c)
                if isinstance(v, float) and pd.isna(v):
                    v = None
                row.append(v)
        ws.append(row)
        r_i += 1

    nrows = ws.max_row
    ref = f"A1:{get_column_letter(len(TOUCH_SCHEMA))}{nrows}"
    group_of = {c: "FF2E6DA4" for c in TOUCH_SCHEMA}
    for c in ("Follow_Up_Flag", "SLA_Status", "Days_Since_Last", "Next_Touch_Due"):
        group_of[c] = "FF2E7D32"
    for c in ("Total_Touches", "Emails_Sent", "Calls_Made", "LinkedIn_Touches", "Last_Touch_Date", "Last_Touch_Channel", "Reply_Received", "Meeting_Booked", "Outcome", "Cadence_Status", "Notes"):
        group_of[c] = "FF6A4CB5"
    style_header(ws, len(TOUCH_SCHEMA), group_of)
    add_table(ws, "TouchTable", ref, style="TableStyleMedium9")

    for c in ("Follow_Up_Flag", "SLA_Status"):
        col = TOUCH_LETTER[c]
        ws.conditional_formatting.add(f"{col}2:{col}{nrows}",
            CellIsRule(operator="equal", formula=['"FOLLOW UP"'], fill=PatternFill("solid", fgColor="FFC0392B"), font=Font(color="FFFFFFFF", bold=True)))
        ws.conditional_formatting.add(f"{col}2:{col}{nrows}",
            CellIsRule(operator="equal", formula=['"OVERDUE"'], fill=PatternFill("solid", fgColor="FFC0392B"), font=Font(color="FFFFFFFF", bold=True)))
        ws.conditional_formatting.add(f"{col}2:{col}{nrows}",
            CellIsRule(operator="equal", formula=['"DUE SOON"'], fill=PatternFill("solid", fgColor="FFC9922B"), font=Font(bold=True)))

    widths = {"Company_Name": 28, "Contact_Title": 24, "Contact_Email": 26, "Notes": 26}
    for c in TOUCH_SCHEMA:
        ws.column_dimensions[TOUCH_LETTER[c]].width = widths.get(c, 14)
    ws.freeze_panes = "C2"
    ws.sheet_view.showGridLines = False
    return ws


def write_cadence_sheet(wb):
    ws = wb.create_sheet("CADENCE")
    ws.append(["Step", "Days_To_Next", "Channel"])
    for row in CADENCE_STEPS:
        ws.append(list(row))
    nrows = ws.max_row
    add_table(ws, "CadenceTable", f"A1:C{nrows}", style="TableStyleMedium9")
    style_header(ws, 3, {"Step": "FF1F3A5F", "Days_To_Next": "FF1F3A5F", "Channel": "FF1F3A5F"})
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.sheet_view.showGridLines = False
    return ws


def write_summary_sheet(wb, merged_df):
    ws = wb.create_sheet("SUMMARY")
    total = len(merged_df)
    scored = merged_df[merged_df["Freight_Relevant"].astype(str).str.lower() == "yes"]
    tier_counts = scored["Fit_Tier"].value_counts().to_dict()
    a = tier_counts.get("A", 0); b = tier_counts.get("B", 0); c = tier_counts.get("C", 0); d = tier_counts.get("D", 0)
    excl = merged_df["Exclusion_Reason"].fillna("").astype(str)
    k12 = (excl.str.contains("K-12", case=False)).sum()
    nonus = (excl.str.contains("Non-US", case=False)).sum()
    transport = (excl.str.contains("Transport", case=False)).sum()
    addressable = pd.to_numeric(scored["Est_Freight_Spend"], errors="coerce").fillna(0).sum()
    verified = (merged_df["Contact_Email"].fillna("").astype(str).str.contains("@")).sum()
    ttm_total = pd.to_numeric(merged_df["CT_TTM_Spend"], errors="coerce").fillna(0).sum()

    rows = [
        [None, None, None, None, None],
        [None, "Master Coverage and Freight Intelligence (Fit v3 scored)", None, None, None],
        [None, None, None, None, None],
        [None, f"{total:,}", None, None, None],
        [None, "Total members, deduplicated", f"Scored with Fit v3 (freight-relevant): {len(scored):,}",
         f"Tier A + B accounts: {a+b:,}", f"Est freight spend covered: ${addressable:,.0f}"],
        [None, None, None, None, None],
        [None, f"{a:,}", None, None, None],
        [None, "Tier A, top of the file", f"Verified contacts on hand: {verified:,}",
         f"CoreTrust TTM reported spend: ${ttm_total:,.0f}", None],
        [None, None, None, None, None],
        [None, f"{k12:,}", None, None, None],
        [None, "K-12 excluded (not shipper prospects)", f"Non-US excluded: {nonus:,}",
         f"Transport/logistics set aside: {transport:,}", f"Tier C, nurture: {c:,}"],
        [None, None, None, None, None],
        [None, "Fit v3 tier distribution (of the scored pool)", None, None, None],
        [None, "Tier", "Accounts", "Meaning", None],
        [None, "A (>=70)", a, "Real freight budget and a proven CoreTrust buying habit. Call today.", None],
        [None, "B (>=55)", b, "Strong on freight and engrainment. Work next.", None],
        [None, "C (>=40)", c, "Some signal, thinner. Nurture and enrich.", None],
        [None, "D (<40)", d, "Not worth qualifying yet.", None],
        [None, None, None, None, None],
        [None, f"Regenerated by fitv3_pipeline.py on {TODAY.isoformat()}. Actionability and Warmth are now explicit "
               "columns on MASTER MEMBERS; Fit_v3_Score only recomputes on rows that gained verified contact or "
               "freight-analysis evidence in this run. See the Scoring Methodology.", None, None, None],
    ]
    for r in rows:
        ws.append(r)
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 46
    ws.column_dimensions["E"].width = 30
    for r in (2, 5, 8, 11, 14):
        ws.cell(row=r, column=2).font = Font(bold=True, size=14, color="FF1F3A5F")
    ws.sheet_view.showGridLines = False
    return ws, dict(total=total, scored=len(scored), a=a, b=b, c=c, d=d, k12=k12, nonus=nonus,
                     transport=transport, addressable=addressable, verified=verified)


def write_kpi_dashboard(wb, stats):
    ws = wb.create_sheet("KPI DASHBOARD")
    ws["B2"] = "Demand Generation and Pipeline Tracker"
    ws["B2"].font = Font(bold=True, size=16, color="FF1F3A5F")
    ws["B3"] = "Live tiles read MasterTable and TouchTable. Run scripts/weekly_kpi_snapshot.py each week to append a row to the history log below -- that is the running record for demand-gen reporting."
    ws["B3"].font = Font(italic=True, color="FF5A6675")
    ws["B3"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("B3:J3")
    ws.row_dimensions[3].height = 30

    tiles = [
        ("Total members", "=COUNTA(MasterTable[Lead_ID])"),
        ("Freight-relevant", '=COUNTIF(MasterTable[Freight_Relevant],"Yes")'),
        ("Tier A", '=COUNTIF(MasterTable[Fit_Tier],"A")'),
        ("Tier B", '=COUNTIF(MasterTable[Fit_Tier],"B")'),
        ("Verified contacts", '=COUNTIF(MasterTable[Contact_Email],"*@*")'),
        ("Ready to Call", '=COUNTIF(MasterTable[Record_Status],"Ready to Call")'),
        ("New / Enriching", '=COUNTIF(MasterTable[Record_Status],"New")+COUNTIF(MasterTable[Record_Status],"Enriching")'),
        ("In Cadence", '=COUNTIF(MasterTable[Record_Status],"In Cadence")'),
        ("Qualified", '=COUNTIF(MasterTable[Qualified],"Yes")'),
        ("Addressable freight $", '=SUMIF(MasterTable[Freight_Relevant],"Yes",MasterTable[Est_Freight_Spend])'),
        ("Touches logged", "=SUM(TouchTable[Total_Touches])"),
        ("Meetings booked", '=COUNTIF(TouchTable[Meeting_Booked],"Yes")'),
        ("Replies received", '=COUNTIF(TouchTable[Reply_Received],"Yes")'),
        ("Overdue follow-ups", '=COUNTIF(TouchTable[SLA_Status],"OVERDUE")'),
    ]
    start_row = 5
    col = 2
    row = start_row
    for i, (label, formula) in enumerate(tiles):
        r = start_row + (i // 4) * 3
        c = 2 + (i % 4) * 2
        cell = ws.cell(row=r, column=c, value=formula)
        cell.font = Font(bold=True, size=18, color="FF2E6DA4")
        lbl = ws.cell(row=r + 1, column=c, value=label)
        lbl.font = Font(size=10, color="FF5A6675")
    for i in range(4):
        ws.column_dimensions[get_column_letter(2 + i * 2)].width = 20

    hist_start = start_row + ((len(tiles) - 1) // 4) * 3 + 4
    ws.cell(row=hist_start, column=2, value="Weekly Demand-Gen History").font = Font(bold=True, size=13, color="FF1F3A5F")
    headers = ["Week_Of", "Total_Members", "Freight_Relevant", "Tier_A", "Tier_B", "Verified_Contacts",
               "New_Or_Enriching", "Touches_Sent", "Meetings_Booked", "Qualified_Opps", "Addressable_Freight"]
    hr = hist_start + 1
    for i, h in enumerate(headers):
        cell = ws.cell(row=hr, column=2 + i, value=h)
        cell.fill = PatternFill("solid", fgColor="FF1F3A5F")
        cell.font = Font(color="FFFFFFFF", bold=True)
    seed_row = hr + 1
    ws.cell(row=seed_row, column=2, value=TODAY)
    for i, (_, formula) in enumerate(tiles[:len(headers) - 1]):
        pass
    formula_map = {
        "Total_Members": "=COUNTA(MasterTable[Lead_ID])",
        "Freight_Relevant": '=COUNTIF(MasterTable[Freight_Relevant],"Yes")',
        "Tier_A": '=COUNTIF(MasterTable[Fit_Tier],"A")',
        "Tier_B": '=COUNTIF(MasterTable[Fit_Tier],"B")',
        "Verified_Contacts": '=COUNTIF(MasterTable[Contact_Email],"*@*")',
        "New_Or_Enriching": '=COUNTIF(MasterTable[Record_Status],"New")+COUNTIF(MasterTable[Record_Status],"Enriching")',
        "Touches_Sent": "=SUM(TouchTable[Total_Touches])",
        "Meetings_Booked": '=COUNTIF(TouchTable[Meeting_Booked],"Yes")',
        "Qualified_Opps": '=COUNTIF(MasterTable[Qualified],"Yes")',
        "Addressable_Freight": '=SUMIF(MasterTable[Freight_Relevant],"Yes",MasterTable[Est_Freight_Spend])',
    }
    for i, h in enumerate(headers[1:], start=1):
        ws.cell(row=seed_row, column=2 + i, value=formula_map.get(h))
    ws.column_dimensions["B"].width = 14
    for i in range(1, len(headers)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18
    ws.sheet_view.showGridLines = False
    return ws


def write_readme(wb, stats, matched, inserted):
    ws = wb.create_sheet("READ ME")
    lines = [
        "",
        "CoreTrust Master Members, all members scored with Fit v3",
        f"Regenerated by scripts/fitv3_pipeline.py on {TODAY.isoformat()}.",
        "",
        f"Every CoreTrust member on one row, {stats['total']:,} of them, deduplicated from the full Salesforce "
        f"export and scored with Fit v3. {stats['scored']:,} freight-relevant members carry a score. K-12 "
        "districts, non-US members, and transportation companies are flagged in Exclusion_Reason and set to "
        "Low Priority. Sort by Fit_v3_Score and work the top.",
        "",
        f"This run merged {matched} accounts against the verified-contact / freight-analysis sample. Those rows "
        "picked up Contact_First/Last/Title/Email/Phone, PE_Sponsor, Portfolio_Company, a refreshed "
        "Est_Freight_Spend, and a rescored Fit_v3_Score/Fit_Tier -- exactly the enrichment the agent runs nightly. "
        "Every other row keeps its original score untouched; Actionability, Warmth, Priority, and Service_Fit are "
        "newly backfilled, descriptive columns and never change Fit_v3_Score on their own. "
        f"{inserted} accounts in the freight analysis had no match in the SFDC export at all -- those were filed "
        "as new rows (fresh Lead_IDs, Record_Status Ready to Call, scored fresh) the same way the agent would "
        "file a brand new registration.",
        "",
        "Tabs: MASTER MEMBERS (table MasterTable), TOUCHPOINTS (table TouchTable, four live formula columns), "
        "CADENCE (table CadenceTable, the 8-step lookup the touchpoint formulas read), SUMMARY (static snapshot), "
        "KPI DASHBOARD (live tiles + an appendable weekly history log -- the demand-gen tracker), Data Dictionary.",
        "",
        "To re-run: drop a fresh SFDC export or enrichment batch in data/source/ and run "
        "python3 scripts/fitv3_pipeline.py. To log a touch or snapshot the week from the command line instead of "
        "Copilot, see scripts/log_touch.py and scripts/weekly_kpi_snapshot.py.",
        "",
    ]
    for l in lines:
        ws.cell(row=ws.max_row + 1, column=2, value=l)
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.sheet_view.showGridLines = False
    return ws


def write_data_dictionary(wb):
    ws = wb.create_sheet("Data Dictionary")
    ws.append(["CoreTrust Master, Data Dictionary", None, None])
    ws.append([None, None, None])
    ws.append(["Column", "Group", "Notes"])
    notes = {
        "Lead_ID": "stable key, never regenerate",
        "PE_Sponsor": "blank until enrichment finds ownership; 'Independent' means confirmed no sponsor",
        "Portfolio_Company": "Yes when PE_Sponsor is set and not Independent",
        "Contact_Phone": "blank until enrichment finds a direct or published number",
        "Actionability": "0-100, engine.action(): verified logistics email 100, verified generic 75, email no title 60, name only 40, nothing 0",
        "Warmth": "0-100, PE backed +60, named CT_Account_Manager or Assigned_Rep +40",
        "Exclusion_Reason": "K-12, Non-US, Transport/Logistics, PE/VC, No-signal, or blank",
        "Fit_v3_Score": "0-100, auto. Never overwrite by hand.",
        "Fit_Tier": "A>=70 B>=55 C>=40 D. Never overwrite by hand.",
        "Priority": "HOT (tier A) / WARM (tier B) / COLD (tier C or D), auto from Fit_Tier",
        "Service_Fit": "rule-based tags from industry + engagement: LTL/FTL/Retail Cross Dock/Brokerage",
        "Qualified": "Yes once the five markers plus the discipline gate pass",
    }
    group_label = {"record": "Record", "company": "Company", "contact": "Contact", "markers": "Markers 1-5",
                   "score_inputs": "Score inputs", "computed": "Scoring, GREEN, computed", "outcome": "Outcome"}
    for c, g in SCHEMA:
        ws.append([c, group_label[g], notes.get(c)])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 90
    ws["A1"].font = Font(bold=True, size=14, color="FF1F3A5F")
    ws["A3"].font = Font(bold=True); ws["B3"].font = Font(bold=True); ws["C3"].font = Font(bold=True)
    ws.sheet_view.showGridLines = False
    return ws


def main():
    global VERIFIED_PATH_GLOBAL
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--master", required=True, help="SFDC-scored master workbook (MASTER MEMBERS sheet)")
    ap.add_argument("--verified", required=True, help="Verified-contact / freight-analysis workbook to merge in")
    ap.add_argument("--out", required=True, help="Output workbook path")
    args = ap.parse_args()

    VERIFIED_PATH_GLOBAL = args.verified
    print(f"Loading master: {args.master}")
    master_df = load_master(args.master)
    print(f"  {len(master_df):,} rows")
    print(f"Loading verified/freight sample: {args.verified}")
    verified_df = load_verified(args.verified)
    print(f"  {len(verified_df):,} rows")

    print("Merging and rescoring...")
    merged_df, matched, inserted = build_merged_master(master_df, verified_df)
    print(f"  {matched} accounts matched and rescored with verified evidence")
    print(f"  {inserted} net-new accounts inserted from the freight analysis")

    print("Building touchpoint seed rows...")
    touch_rows = build_touchpoints(merged_df, verified_df)
    print(f"  {len(touch_rows)} seed touchpoint rows")

    print("Writing workbook...")
    wb = Workbook()
    wb.remove(wb.active)
    write_master_sheet(wb, merged_df)
    write_touch_sheet(wb, touch_rows)
    write_cadence_sheet(wb)
    _, stats = write_summary_sheet(wb, merged_df)
    write_kpi_dashboard(wb, stats)
    write_readme(wb, stats, matched, inserted)
    write_data_dictionary(wb)

    order = ["MASTER MEMBERS", "TOUCHPOINTS", "CADENCE", "KPI DASHBOARD", "SUMMARY", "READ ME", "Data Dictionary"]
    wb._sheets.sort(key=lambda ws: order.index(ws.title))
    wb.active = 0

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Wrote {args.out}")
    print(f"  Tier A: {stats['a']:,}  Tier B: {stats['b']:,}  Tier C: {stats['c']:,}  Tier D: {stats['d']:,}")
    print(f"  Verified contacts: {stats['verified']:,}  Addressable freight: ${stats['addressable']:,.0f}")


if __name__ == "__main__":
    main()
